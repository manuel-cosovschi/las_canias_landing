"""Extrae cualquier temporada, encontrando los bloques de casa por el encabezado
'Mes' en vez de tenerlos hardcodeados como extraer.py."""
import openpyxl, json, sys

COLS = ['mes','importe','dias','prom','fecha','anticipo','restan',
        'inquilinos','telefono','dni','fac','dolar','dolares']

def limpio(v):
    if v is None: return None
    s = str(v).strip()
    return s or None

def bloques(ws):
    """Filas donde arranca cada casa. El complejo fue creciendo, así que cada
    temporada tiene una cantidad distinta de bloques."""
    hits = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == 'mes':
                hits.append((r, c))
    return hits

def extraer_hoja(ws):
    hits = bloques(ws)
    casas = {}
    for i, (fila_enc, col) in enumerate(hits):
        fin = hits[i+1][0] - 2 if i + 1 < len(hits) else ws.max_row
        filas = []
        for r in range(fila_enc + 1, fin + 1):
            fila = {k: limpio(ws.cell(r, col + off).value) for off, k in enumerate(COLS)}
            if any(fila.values()):
                fila['_fila'] = r
                filas.append(fila)
        casas[f'LC{i+1}'] = filas
    return casas

if __name__ == '__main__':
    wb = openpyxl.load_workbook('alquileres.xlsx', data_only=True)
    salida = {t: extraer_hoja(wb[t]) for t in sys.argv[1:]}
    json.dump(salida, open('extraido_hist.json','w'), ensure_ascii=False, indent=1)
    for t, casas in salida.items():
        print(f'{t}: ' + ', '.join(f'{c}={len(f)}' for c, f in casas.items()))
