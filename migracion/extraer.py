import openpyxl, json, re
from openpyxl.utils import get_column_letter as L

wb = openpyxl.load_workbook('alquileres.xlsx', data_only=True)

# La tabla arranca en AH; los encabezados se repiten por casa
BLOQUES = [(5, 'LC1'), (46, 'LC2'), (88, 'LC3'), (129, 'LC4'), (167, 'LC5')]
COLS = {  # offset desde AH (col 34)
    'mes': 0, 'importe': 1, 'dias': 2, 'prom': 3, 'fecha': 4, 'anticipo': 5,
    'restan': 6, 'inquilinos': 7, 'telefono': 8, 'dni': 9, 'fac': 10,
    'dolar': 11, 'dolares': 12,
}
AH = 34

def limpio(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def extraer(ws, fila_ini, fila_fin):
    filas = []
    for r in range(fila_ini, fila_fin + 1):
        fila = {k: limpio(ws.cell(r, AH + off).value) for k, off in COLS.items()}
        # extras a la derecha de 'dolares'
        extra = [limpio(ws.cell(r, c).value) for c in range(AH + 13, ws.max_column + 1)]
        extra = [e for e in extra if e]
        if any(fila.values()) or extra:
            fila['_fila'] = r
            if extra:
                fila['_extra'] = extra
            filas.append(fila)
    return filas

salida = {}
for nombre in ['Temporada 25-26', 'Temporada 26-27']:
    ws = wb[nombre]
    casas = {}
    for i, (fila_enc, casa) in enumerate(BLOQUES):
        fin = BLOQUES[i + 1][0] - 2 if i + 1 < len(BLOQUES) else ws.max_row
        casas[casa] = extraer(ws, fila_enc + 1, fin)
    salida[nombre] = casas

json.dump(salida, open('extraido.json', 'w'), ensure_ascii=False, indent=1)

for temp, casas in salida.items():
    print('=' * 72)
    print(temp)
    total_filas = sum(len(v) for v in casas.values())
    print(f'  filas con algo: {total_filas}')
    for casa, filas in casas.items():
        con_nombre = [f for f in filas if f['inquilinos']]
        print(f'  {casa}: {len(filas)} filas, {len(con_nombre)} con inquilino')
